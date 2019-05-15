import time
from openpyxl import load_workbook
from netmiko import ConnectHandler, redispatch
import os
from getpass import getpass

wb = load_workbook(filename='Planilla_Solicitud_de_Interfaces_TK3164_act.xlsx', read_only=True)
ws = wb['PLANILLA']
sol=[]
aux = []
aux2 = []
intf = []

user = input('Enter OpManager username: ')
passwd = getpass('Enter OpManager password: ')
ustac = input('Enter OpManager username: ')
passtac = getpass('Enter OpManager password: ') 

jumpserver = {'device_type': 'terminal_server',
    'ip': '192.168.97.196',
    'username': user,
    'password': passwd,
    'global_delay_factor':1}

for row in ws['B29':'U100']:
	for cell in row:
		if '------' in str(cell.value) or 'Justifica' in str(cell.value): 
			break
		if (cell.value != None):
			if '.M' in str(cell):
				sol.append(cell.value)
			if '.B' in str(cell):
				aux.append(cell.value)
			if '.C' in str(cell):
				aux2.append(cell.value)
			if '.U' in str(cell):
				intf.append(cell.value)		
	else:
		continue
	break

for i in range(len(aux)):
	aux[i] = aux[i].replace(' ', '_') + '_' + aux2[i].replace(' ', '_')

for y in range(len(sol)):
	#find and store index of repeated devices
	indices = [i for i, x in enumerate(sol) if x == sol[y]]
	print (indices)
	#compares if the device read from file is not repeated
	if y == indices[0]:
		net_connect = ConnectHandler(**jumpserver)
		print (net_connect.find_prompt())
		print ('connecting to ' + sol[y])
		net_connect.write_channel('ssh ' + ustac + '@' + sol[y])
		max_loops = 10
		i = 1
		while i <= max_loops:
			output = net_connect.read_channel()
			# Search for password pattern / send password
			if 'assword' in output:
				net_connect.write_channel(passtac + '\r\n')
				time.sleep(.5)
				output = net_connect.read_channel()
				# Did we successfully login
				if '>' in output or '#' in output:
					break

			net_connect.write_channel('\r\n')
			time.sleep(1)
			i += 1

		net_connect.read_channel()

		redispatch(net_connect, device_type='cisco_ios')
		
		#applies configuration to device interfaces
		for each in indices:
			output = net_connect.send_config_set(['int eth 1/'+str(intf[each]),'desc '+aux[each] + '(Reservada)'], exit_config_mode = False)
			print (output)
		print (net_connect.send_command('show int desc'))
		net_connect.disconnect()

