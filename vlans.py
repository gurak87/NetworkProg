from getpass import getpass
from jumpssh import SSHSession
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill
import datetime
#import time
import re

file = open('ciscoip.txt')
wb = Workbook()
ws = wb.active
ws.tittle = 'Vlans'
ws['A1'] = 'Hostname' 
ws['B1'] = 'VLAN'
ws['C1'] = 'Name'
index = 0
fill = PatternFill("solid", fgColor="DDDDDD")


#Fill header row using bold font and fill color
for row in ws.iter_rows(min_row=1, max_col=3, max_row=1):
	for cell in row:
		cell.font = Font(bold=True)
		cell.fill = fill

user = input('Enter OpManager username: ')
passwd = getpass('Enter OpManager password: ')
		
#Open connection to GW server
gw = SSHSession('192.168.97.196', user, password=passwd).open()

user = input('Enter TACACS username: ')
passwd = getpass('Enter TACACS password: ')

#Adds info from a list to a column in excel file
def agregate (lt, y):
	for each in range(len(y)):
		ws[lt+str(each+2+index)] = y[each]
	
count = 1

def RepresentsInt(s):
    try: 
        int(s)
        return True
    except ValueError:
        return False

for device in file:
	aux = []
	aux2 = []
	vlan = []
	result = ''
	#Open connection to remote device
	remote = gw.get_remote_session(device, username=user, password=passwd)
	
	result = remote.get_cmd_output('show vlan brief').split()

	for each in result:
		if RepresentsInt(each):
			vlan.append(result[result.index(each)])
			aux.append(result[result.index(each)+1])		
	agregate('B',vlan)
	agregate('C',aux)
	aux = []
	
#Adds hostname
	aux = remote.get_cmd_output('show hostname')
	for each in range(len(vlan)):
		ws['A'+str(each+2+index)] = aux

	print (count)
	count += 1

	remote.close()
	index += len(vlan)

wb.save('vlans'+str(datetime.date.today())+'.xlsx')
gw.close()

