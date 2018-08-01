from getpass import getpass
from jumpssh import SSHSession
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill
import datetime
import re

file = open('test7k.txt')
wb = Workbook()
ws = wb.active
ws.tittle = 'Modulo'
ws['A1'] = 'Hostname' 
ws['B1'] = 'Interface'
ws['C1'] = 'Type'
ws['D1'] = 'Serial'
ws['E1'] = 'Status'
ws['F1'] = 'Speed'
ws['G1'] = 'Description'
index = 0
fill = PatternFill("solid", fgColor="DDDDDD")

user = input('Enter GW server username: ')
passwd = getpass('Enter GW server password: ')

for row in ws.iter_rows(min_row=1, max_col=7, max_row=1):
	for cell in row:
		cell.font = Font(bold=True)
		cell.fill = fill
#Open connection to GW server
gw = SSHSession('192.168.97.196', user, password=passwd).open()

user = input('Enter TACACS username: ')
passwd = getpass('Enter TACACS password: ')

for device in file:
	aux1 = []
	aux2 = []
	poinsw = []
	result = ''
	#Open connection to remote device
	remote = gw.get_remote_session(device, username=user, password=passwd)

	def agregate (lt, y):
		for each in range(len(y)):
			ws[lt+str(each+2+index)] = y[each]
		
	def separate (cmd,aux):
		aux = remote.get_cmd_output(cmd).split()
		return aux
	
	result = separate('show int status | in Eth',result)

#Check all interfaces and current status (up/down)
	for each in range(len(result)):
		if re.match(r'(Te|Eth)\d{1,3}\/\d{1,2}(\/\d)*',result[each]):
			poinsw.append(result[each])
		if result[each] == 'connected':
			aux2.append('up')
		elif re.match(r'.*(not|disa|bsen|nelDo|FlapE|suspnd).*',result[each]):
			aux2.append('down')

	agregate('B',poinsw)
	agregate('E',aux2)
	aux2 = []
	
	result = separate('show int desc | in Eth',result)

#Check speed and description per interface
	for each in range(len(result)):
		if result[each] == 'eth':
			aux1.append(result[each+1])
			staux = ''
			for aux in result[each+2:]:
				if re.match(r'Eth\d{1,3}\/\d{1,2}(\/\d)*',aux):
					break
				elif re.match(r'.*[L,l]ink',aux):
					staux += aux+' '
					break
				else:
					staux += aux+' '
			aux2.append(staux)

	agregate('F',aux1)
	agregate('G',aux2)

#Adds hostname
	aux2 = remote.get_cmd_output('show hostname')
	for each in range(len(poinsw)):
		ws['A'+str(each+2+index)] = aux2

	aux1 = []
	aux2 = []

#closes remote session and reopens it
#done because of problems with some devices closing the session before getting all the ouput
	remote.close()
	remote = gw.get_remote_session(device, username=user, password=passwd)

#Checks for SFP modules
	result = separate('show interface transceiver',result)
	
	for i in range(len(result)-1):
		if result[i] == 'not':
			aux1.append('--')
			aux2.append('--')
		elif result[i] == 'serial':
			aux1.append(result[i+3])
		elif result[i] == 'type':
			aux2.append(result[i+2])

	agregate('C',aux2)
	agregate('D',aux1)
	remote.close()
	index += len(poinsw)
	
wb.save('modulos'+str(datetime.date.today())+'.xlsx')
gw.close()